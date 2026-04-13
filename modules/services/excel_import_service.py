"""
ExcelImportService: Excel file parsing and import logic
Handles reading Excel files and creating quizzes from them
"""

import openpyxl
from io import BytesIO
from ..models import db, Quiz, Question
from .validation_service import ValidationService
from .quiz_service import QuizService


class ExcelImportService:
    """Service for importing quizzes from Excel files - Instance-based for better performance"""

    # Excel sheet structure constants
    QUESTION_COLUMN = 1      # Column A: Question text
    OPTION_A_COLUMN = 2      # Column B: Option A
    OPTION_B_COLUMN = 3      # Column C: Option B
    OPTION_C_COLUMN = 4      # Column D: Option C
    OPTION_D_COLUMN = 5      # Column E: Option D
    CORRECT_ANSWER_COLUMN = 6  # Column F: Correct answer (A/B/C/D)
    DIFFICULTY_COLUMN = 7    # Column G: Difficulty (1-5, optional)
    
    DATA_START_ROW = 2       # Data starts from row 2 (row 1 is header)
    
    EXPECTED_HEADERS = [
        'Câu Hỏi',
        'Đáp án A',
        'Đáp án B',
        'Đáp án C',
        'Đáp án D',
        'Đáp án Đúng',
        'Mức Độ Khó'
    ]

    def __init__(self, file_content):
        """
        Initialize ExcelImportService with file content
        Load and validate Excel file once
        
        Args:
            file_content (bytes) - Excel file content
        """
        self.file_content = file_content
        self.workbook = None
        self.worksheet = None
        self.questions = None
        self.error_message = None
        self.is_valid = False
        
        # Try to load workbook
        try:
            self.workbook = openpyxl.load_workbook(BytesIO(file_content))
        except openpyxl.utils.exceptions.InvalidFileException:
            self.error_message = "Invalid Excel file format"
        except Exception as e:
            self.error_message = f"Failed to read Excel file: {str(e)}"

    def validate(self):
        """
        Validate Excel file format and structure
        Returns:
            bool - True if valid, False if not
        """
        try:
            # Check if workbook loaded successfully
            if not self.workbook:
                return False
            
            # Check if 'questions' sheet exists
            if 'questions' not in self.workbook.sheetnames:
                self.error_message = "Excel file must contain a 'questions' sheet"
                return False
            
            self.worksheet = self.workbook['questions']
            
            # Check headers (Vietnamese format)
            for idx, expected_header in enumerate(self.EXPECTED_HEADERS, 1):
                cell_value = self.worksheet.cell(1, idx).value
                if not cell_value or str(cell_value).strip() != expected_header:
                    self.error_message = f"Column {chr(64+idx)} header should be '{expected_header}'"
                    return False
            
            # Check for at least one data row
            if self.worksheet.max_row < 2:
                self.error_message = "Excel file must contain at least one question"
                return False
            
            self.is_valid = True
            return True
        
        except Exception as e:
            self.error_message = f"Error validating Excel file: {str(e)}"
            return False

    def parse(self, quiz_name):
        """
        Parse Excel file and extract questions from validated worksheet
        
        Args:
            quiz_name (str) - Name for the quiz (used for error messages)
        
        Returns:
            dict - {
                'success': bool,
                'data': questions_list or None,
                'message': str
            }
        """
        # Check if validation was successful
        if not self.is_valid:
            return {
                'success': False,
                'data': None,
                'message': self.error_message or 'File not validated'
            }
        
        try:
            self.questions = []
            
            # Read data rows
            for row_idx in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
                # Get cell values
                question_text = self.worksheet.cell(row_idx, self.QUESTION_COLUMN).value
                option_a = self.worksheet.cell(row_idx, self.OPTION_A_COLUMN).value
                option_b = self.worksheet.cell(row_idx, self.OPTION_B_COLUMN).value
                option_c = self.worksheet.cell(row_idx, self.OPTION_C_COLUMN).value
                option_d = self.worksheet.cell(row_idx, self.OPTION_D_COLUMN).value
                correct_answer = self.worksheet.cell(row_idx, self.CORRECT_ANSWER_COLUMN).value
                difficulty = self.worksheet.cell(row_idx, self.DIFFICULTY_COLUMN).value
                
                # Skip empty rows
                if not question_text:
                    continue
                
                # Validate and create question
                question_data = {
                    'question_text': str(question_text).strip(),
                    'option_a': str(option_a).strip() if option_a else '',
                    'option_b': str(option_b).strip() if option_b else '',
                    'option_c': str(option_c).strip() if option_c else '',
                    'option_d': str(option_d).strip() if option_d else '',
                    'correct_answer': str(correct_answer).upper().strip() if correct_answer else 'A',
                    'difficulty': int(difficulty) if difficulty else 3
                }
                
                # Validate question
                is_valid, error_msg = ValidationService.validate_question_format(question_data)
                if not is_valid:
                    self.error_message = f"Error in row {row_idx}: {error_msg}"
                    return {
                        'success': False,
                        'data': None,
                        'message': self.error_message
                    }
                
                self.questions.append(question_data)
            
            if not self.questions:
                self.error_message = "No valid questions found in Excel file"
                return {
                    'success': False,
                    'data': None,
                    'message': self.error_message
                }
            
            return {
                'success': True,
                'data': self.questions,
                'message': f"Parsed {len(self.questions)} questions from Excel file"
            }
        
        except Exception as e:
            self.error_message = f"Error parsing Excel file: {str(e)}"
            return {
                'success': False,
                'data': None,
                'message': self.error_message
            }

    def import_as_quiz(self, quiz_name):
        """
        Create quiz from parsed questions
        
        Args:
            quiz_name (str) - Name for the new quiz
        
        Returns:
            dict - {
                'success': bool,
                'data': quiz_obj or None,
                'message': str
            }
        """
        if not self.questions:
            return {
                'success': False,
                'data': None,
                'message': 'No questions parsed yet. Call parse() first'
            }
        
        # Create quiz using QuizService
        result = QuizService.create_quiz(quiz_name, self.questions)
        return result

    @staticmethod
    def get_import_template():
        """
        Generate an Excel template for quiz import (static - no file dependency)
        
        Returns:
            BytesIO - Excel file template
        """
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'questions'
            
            # Create headers
            from openpyxl.styles import Font
            for col_idx, header in enumerate(ExcelImportService.EXPECTED_HEADERS, 1):
                cell = worksheet.cell(1, col_idx)
                cell.value = header
                cell.font = Font(bold=True)
            
            # Add sample row
            sample_data = [
                'Thủ đô của Việt Nam là gì?',
                'Hà Nội',
                'Hồ Chí Minh',
                'Hải Phòng',
                'Đà Nẵng',
                'A',
                '1'
            ]
            
            for col_idx, value in enumerate(sample_data, 1):
                worksheet.cell(2, col_idx).value = value
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 20
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            
            # Write to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output
        
        except Exception as e:
            return None
